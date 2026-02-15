"""Export assessments to multi-sheet Excel workbook.

Sheets:
1. Summary — Study, D1-D5 judgments, Overall, Confidence
2. Per-domain detail — Q&A with quotes and justifications
3. Metadata — Model, timestamps, versions
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from rob_assess.models.assessment import StudyAssessment
from rob_assess.models.common import RoB2Domain, RoB2Judgment

DOMAIN_ORDER = [RoB2Domain.D1, RoB2Domain.D2, RoB2Domain.D3, RoB2Domain.D4, RoB2Domain.D5]

# Color scheme matching robvis/Cochrane
JUDGMENT_FILLS = {
    RoB2Judgment.LOW: PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    RoB2Judgment.SOME_CONCERNS: PatternFill(
        start_color="FFC000", end_color="FFC000", fill_type="solid"
    ),
    RoB2Judgment.HIGH: PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    ),
}

HEADER_FONT = Font(bold=True, size=11)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def export_excel(
    assessments: list[StudyAssessment],
    output_path: str | Path,
) -> Path:
    """Export assessments to a multi-sheet Excel workbook."""
    output_path = Path(output_path)
    wb = openpyxl.Workbook()

    _write_summary_sheet(wb.active, assessments)
    _write_detail_sheet(wb.create_sheet("Domain Details"), assessments)
    _write_metadata_sheet(wb.create_sheet("Metadata"), assessments)

    wb.save(str(output_path))
    return output_path


def _write_summary_sheet(ws, assessments: list[StudyAssessment]):
    """Summary sheet with traffic-light colored judgments."""
    ws.title = "Summary"
    headers = ["Study", "D1", "D2", "D3", "D4", "D5", "Overall", "Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT

    for row_idx, sa in enumerate(assessments, 2):
        ws.cell(row=row_idx, column=1, value=sa.study_id)

        if sa.assessment:
            for col_idx, domain in enumerate(DOMAIN_ORDER, 2):
                da = sa.assessment.domains.get(domain)
                if da and da.judgment:
                    cell = ws.cell(row=row_idx, column=col_idx, value=da.judgment.value)
                    fill = JUDGMENT_FILLS.get(da.judgment)
                    if fill:
                        cell.fill = fill

            overall = sa.assessment.overall_judgment
            if overall:
                cell = ws.cell(row=row_idx, column=7, value=overall.value)
                fill = JUDGMENT_FILLS.get(overall)
                if fill:
                    cell.fill = fill

        ws.cell(row=row_idx, column=8, value=f"{sa.overall_confidence:.0%}")

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)


def _write_detail_sheet(ws, assessments: list[StudyAssessment]):
    """Per-domain signalling questions, answers, and quotes."""
    headers = ["Study", "Domain", "Question #", "Question", "Answer", "Justification", "Quotes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT

    row = 2
    for sa in assessments:
        if not sa.assessment:
            continue
        for domain in DOMAIN_ORDER:
            da = sa.assessment.domains.get(domain)
            if not da:
                continue
            for sq in da.signalling_questions:
                ws.cell(row=row, column=1, value=sa.study_id)
                ws.cell(row=row, column=2, value=domain.value)
                ws.cell(row=row, column=3, value=sq.number)
                ws.cell(row=row, column=4, value=sq.text).alignment = WRAP_ALIGNMENT
                ws.cell(row=row, column=5, value=sq.answer.value if sq.answer else "")
                ws.cell(row=row, column=6, value=sq.justification).alignment = WRAP_ALIGNMENT
                ws.cell(
                    row=row, column=7, value="\n".join(sq.quotes)
                ).alignment = WRAP_ALIGNMENT
                row += 1

    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 40


def _write_metadata_sheet(ws, assessments: list[StudyAssessment]):
    """Metadata about the assessment run."""
    headers = ["Study", "PDF Path", "Model", "Assessed At", "Tool Version", "Overall Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT

    for row_idx, sa in enumerate(assessments, 2):
        ws.cell(row=row_idx, column=1, value=sa.study_id)
        ws.cell(row=row_idx, column=2, value=sa.pdf_path)
        ws.cell(row=row_idx, column=3, value=sa.model_used)
        ws.cell(row=row_idx, column=4, value=sa.assessed_at)
        ws.cell(row=row_idx, column=5, value=sa.tool_version)
        ws.cell(row=row_idx, column=6, value=f"{sa.overall_confidence:.0%}")
